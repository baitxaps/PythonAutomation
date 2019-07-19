#! /usr/bin/env python3
# coding: utf-8
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pprint
from openpyxl.styles import Font  # 设置单元格的字体风格
#import os
#import census2010


def loadxlsxTest():
    wb = openpyxl.load_workbook('./automate_online-materials/example.xlsx')
    print(type(wb))
    print(wb.get_sheet_names())

    # 每个表由一个 Worksheet 对象表示，可以通过向工作簿方法 get_sheet_by_name()传递表名字符串获得
    sheet = wb.get_sheet_by_name('Sheet1')
    print(sheet)

    # 可以通过 title 属性取得它的名称
    print(sheet.title)
    # 取得工作簿的活动表
    antherSheet = wb.get_active_sheet()
    print(antherSheet)

    print(sheet['A1'])
    print(sheet['A1'].value)

    c = sheet['B1']
    print(c.value)
    print('Row ' + str(c.row) + ', Column ' + ' is  ' + str(c.column) + str(c.value))
    print('Cell ' + str(c.coordinate) + ' is ' + str(c.value))

    # 可以通过 Worksheet 对象的 get_highest_row()和 get_highest_column()方法，确定 表的大小
    # get_highest_row()和get_highest_column()在最新版的openpyxl模块中已经被删除了，取而代之的是max_row和max_column两个方法
    rows = sheet.max_row
    colunms = sheet.max_column
    print(colunms)
    for i in range(1, rows, 2):
        print(i, sheet.cell(row=i, column=2).value)

    print(get_column_letter(900))
    print(get_column_letter(colunms))
    print(column_index_from_string('AA'))

    # 取得电子表格中一行、一列或一个矩形区域中的所有 Cell 对象
    print(tuple(sheet['A1':'C3']))
    for rowOfCellObjects in sheet['A1':'C3']:
        for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)
        print('--- END OF ROW ---')


#loadxlsxTest()


# 创建并保存 Excel 文档
def writeExcelDate():
    wb = openpyxl.Workbook()
    print(wb.sheetnames)
    sheet = wb.active  # wb.get_active_sheet()
    sheet.title = 'Spam Bacon eggs sheet'
    print(sheet)

    wb = openpyxl.load_workbook('./automate_online-materials/example.xlsx')
    sheet = wb.active
    sheet.title = 'Spam Spam Spam'
    wb.save('example_copy.xlsx')
    wb.create_sheet(index=0, title='First Sheet')


#writeExcelDate()


def createAndremoveExcelSheet():
    wb = openpyxl.Workbook()
    # print(wb.get_sheet_names())
    print(wb.sheetnames)
    wb.create_sheet()
    print(wb.sheetnames)
    wb.create_sheet(index=0, title='First Sheet')
    print(wb.sheetnames)
    wb.create_sheet(index=2, title='Middle Sheet')
    print(wb.sheetnames)

    # 创建和删除工作表
    wb.remove(wb['Middle Sheet'])
    print(wb.sheetnames)
    del wb['Sheet1']  # wb.get_sheet_by_name('Sheet1')
    print(wb.sheetnames)

    # 将值写入单元格
    sheet = wb.get_sheet_by_name('Sheet')
    sheet['A1'] = 'Hello world!'
    print(sheet['A1'].value)


#createAndremoveExcelSheet()


# 设置单元格的字体风格
def setExcelFont():
    wb = openpyxl.Workbook()
    # sheet = wb.get_sheet_by_name('Sheet')
    sheet = wb['Sheet']
    italic24Font = Font(size=24, italic=True)
    sheet['A1'].font = italic24Font
    sheet['A1'] = 'Hello world!'
    # 公式
    sheet['a1'] = 200
    sheet['a2'] = 300
    sheet['a3'] = '=sum(a1:a2)'
    # 设置行高和列宽
    sheet.row_dimensions[1].height = 70
    sheet.column_dimensions['B'].width = 20
    # 合并和拆分单元格
    sheet.merge_cells('A1:D3')
    sheet['A1'] = 'Twelve cells merged together.'
    sheet.merge_cells('C5:D5')
    sheet['C5'] = 'Two merged cells.'
    # 要拆分单元格
    sheet.unmerge_cells('A1:D3')
    sheet.unmerge_cells('C5:D5')
    # 冻结窗格
    sheet.freeze_panes = 'A2'
    # 图表
    for i in range(1, 11):
        sheet['A' + str(i)] = i
    refObj = openpyxl.chart.Reference(sheet, min_row=1, min_col=1, max_row=10, max_col=1)
    seriesObj = openpyxl.chart.Series(refObj, title='First series')
    chartObj = openpyxl.chart.BarChart()
    chartObj.append(seriesObj)
    chartObj.title = 'My Chart'
    chartObj.append(seriesObj)
    sheet.add_chart(chartObj, 'C5')
    wb.save('styled.xlsx')


#setExcelFont()

'''
project
读取电子表格数据
'''


def readExceldata():
    print('Opening workbook...')
    wb = openpyxl.load_workbook('./automate_online-materials/censuspopdata.xlsx')
    sheet = wb.get_sheet_by_name('Population by Census Tract')

    '''
    {'AK': {'Aleutians East': {'pop': 3141, 'tracts': 1},
           'Aleutians West': {'pop': 5561, 'tracts': 2},
           'Anchorage': {'pop': 291826, 'tracts': 55},
           'Bethel': {'pop': 17013, 'tracts': 3},
           'Bristol Bay': {'pop': 997, 'tracts': 1},
            --snip--
    '''
    countyData = {}

    print('Reading rows...')
    for row in range(2, sheet.max_row + 1):
        # each row in the spreadsheet has data for one census tract.
        state = sheet['B' + str(row)].value
        county = sheet['C' + str(row)].value
        pop = sheet['D' + str(row)].value

        countyData.setdefault(state, {})
        countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
        # Each row represents one census tract, so increment by one.
        countyData[state][county]['tracts'] += 1
        # Increase the county pop by the pop in this census tract.
        countyData[state][county]['pop'] += int(pop)

    print('Writing results...')
    resultFile = open('census2010.py', 'w')
    resultFile.write('allData = ' + pprint.pformat(countyData))
    resultFile.close()
    print('Done')


# def exportData():
#     print(census2010.allData['AK']['Anchorage'])
#     anchoragePop = census2010.allData['AK']['Anchorage']['pop']
#     print('The 2010 population of Anchorage was ' + str(anchoragePop))


#readExceldata()
#exportData()


'''
更新一个电子表格
project
'''


def updateExcelSheet():
    wb = openpyxl.load_workbook('./automate_online-materials/produceSales.xlsx')
    # DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
    # sheet = wb.get_sheet_by_name('Sheet')
    sheet = wb['Sheet']
    # The produce types and their updated prices
    PRICE_UPDATES = {'Garlic': 3.07,
                     'Celery': 1.19,
                     'Lemon': 1.27}

    for rowNum in range(2, sheet.max_row):
        produceName = sheet.cell(row=rowNum, column=1).value
        if produceName in PRICE_UPDATES:
            sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

    wb.save('updatedProduceSales.xlsx')


#updateExcelSheet()

# 乘法表
def multiplicationTable(num):
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='9by9')
    print(wb.sheetnames)
    sheet = wb['9by9']
    for i in range(2, num + 2):  # 在第一行填写1~N
        sheet.cell(row=1, column=i).value = i - 1
    for i in range(2, num + 2):  # 在第一列填写1~N
        sheet.cell(row=i, column=1).value = i - 1
    x = openpyxl.utils.get_column_letter(num + 1) + str(num + 1)  #得到乘法表右下角单元格的坐标
    for rowobj in sheet['B2':x]:  # 迭代乘法表每个单元格，填入数据
        for cellobj in rowobj:
            a = sheet.cell(row=cellobj.row, column=1).value
            b = sheet.cell(row=1, column=cellobj.column).value
            #  b = sheet.cell(row=1, column=openpyxl.utils.column_index_from_string(cellobj.column)).value
            cellobj.value = a * b

    wb.save('9*9.xlsx')


#multiplicationTable(9)


# 空行插入程序
def blankRowInserter(file, N, M):
    print('start...')
    wb = openpyxl.load_workbook('./automate_online-materials/example.xlsx')
    sheet = wb['Sheet1']
    copywb = openpyxl.Workbook()
    copysheet = copywb['Sheet']

    # maxrow = get_column_letter(sheet.max_row)
    # maxcol = get_column_letter(sheet.max_column)
    # for rowNum in sheet['a1':x]:
    for rowNum in range(1, sheet.max_row + 1):
        for colNum in range(1, sheet.max_column + 1):
            if rowNum < N:
                value = sheet.cell(row=rowNum, column=colNum).value
                copysheet.cell(row=rowNum, column=colNum).value = value
            else:
                value = sheet.cell(row=rowNum, column=colNum).value
                copysheet.cell(row=rowNum + M, column=colNum).value = value
            # print("row:" + str(rowNum) + 'col:' + str(colNum))
    copywb.save(file)
    print('Done')


#blankRowInserter('insert.xlsx', 3, 3)


# 电子表格单元格翻转程序
def excelrowToColumn():
    print('start...')
    wb = openpyxl.load_workbook('./automate_online-materials/example.xlsx')
    sheet = wb['Sheet1']
    copywb = openpyxl.Workbook()
    copysheet = copywb['Sheet']
    for rowNum in range(1, sheet.max_row + 1):
        for colNum in range(1, sheet.max_column + 1):
            copysheet.cell(row=colNum, column=rowNum).value = sheet.cell(row=rowNum, column=colNum).value 

    copywb.save("rowToColumn.xlsx")
    print('Done')


#excelrowToColumn()


# 文本文件到电子表格
def copytxtToExcel():
    print('start...')
    file = open('./automate_online-materials/guests.txt', 'r')
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    col = 1
    for line in file.readlines():
        print(line)
        sheet.cell(row=1, column=col).value = line
        col = col + 1
    file.close()
    wb.save("copytxtToExcel.xlsx")
    print('Done...')


#copytxtToExcel()


# 电子表格到文本文件
def copyexcelToTxt():
    print('start...')
    file = open('guests.txt', 'wb')
    wb = openpyxl.load_workbook('./automate_online-materials/example.xlsx')
    sheet = wb['Sheet1']

    # getexcel = list(sheet.columns)
    # for i in range(len(getexcel[0])):
    #     value = (str(getexcel[0][i].value)).encode()
    #     print(type(value))
    #     file.write(value)
    for rowNum in range(1, sheet.max_row + 1):
        for colNum in range(1, sheet.max_column + 1):
            cellval = sheet.cell(row=rowNum, column=colNum).value
            file.write(str(cellval).encode())
        file.write(str('\n').encode())
    file.close()
    print('Done...')


copyexcelToTxt()

'''
1.openpyxl.load_workbook()函数返回什么?
2.get_sheet_names()工作簿方法返回什么?
3.如何取得名为'Sheet1'的工作表的 Worksheet 对象?
4.如何取得工作簿的活动工作表的 Worksheet 对象?
5.如何取得单元格 C5 中的值?
6.如何将单元格 C5 中的值设置为"Hello"?
7.如何取得表示单元格的行和列的整数?
8.工作表方法 get_highest_column()和 get_highest_row()返回什么?这些返回值
的类型是什么?
9.如果要取得列'M'的整数下标，需要调用什么函数? openpyxl.cell.column_index_from_string('M')
10.如果要取得列 的字符串名称，需要调用什么函数?
11.如何取得从 A1 到 F1 的所有 Cell 对象的元组? 1
2.如何将工作簿保存到文件名 example.xlsx?
13.如何在一个单元格中设置公式?
14.如果需要取得单元格中公式的结果，而不是公式本身，必须先做什么?
15.如何将第 5 行的高度设置为 100?
16.如何设置列 C 的宽度?
17.列出一些 openpyxl 2.1.4 不会从电子表格文件中加载的功能。 18.什么是冻结窗格?
19.创建一个条形图，需要调用哪 5 个函数和方法?
'''
